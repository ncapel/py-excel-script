import openpyxl
from collections import defaultdict
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def optimize_count_multiple(ws):
    data = defaultdict(lambda: defaultdict(int))
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 32:  # Ensure the row has enough columns
            continue  # Skip this row if it doesn't have enough columns
        
        phrase = row[11] if len(row) > 11 else None  # Assuming phrase is in column L (12th column)
        manually_indexed = row[24] == "Manually Indexed" if len(row) > 24 else False  # Column Y
        indexer_review = row[31] == "Yes" if len(row) > 31 else False  # Column AF
        needs_review = [row[i] == "NEEDSREVIEW" for i in range(5) if len(row) > i]  # Columns A-E
        patient_found = row[4] != "NOTFOUND" if len(row) > 4 else False  # Column E
        provider_needs_review = row[5] == "NEEDSREVIEW" if len(row) > 5 else False  # Column F

        if phrase and manually_indexed and indexer_review:
            data[phrase]['total'] += 1
            for i, val in enumerate(needs_review):
                if val:
                    data[phrase][f'col_{i}'] += 1
            if patient_found and provider_needs_review:
                data[phrase]['provider_changed'] += 1

    return data

def filed_documents_report_with_phrase_hit_athena():
    # Load the workbook
    wb = openpyxl.load_workbook('test_sheet.xlsx')
    ws = wb['Documents Filed Report']

    # Print worksheet information for debugging
    print(f"Number of rows: {ws.max_row}")
    print(f"Number of columns: {ws.max_column}")
    print(f"Column headers: {[cell.value for cell in ws[1]]}")

    # Optimize counting
    count_data = optimize_count_multiple(ws)

    # Add Phrase Maintenance Sheet
    phrase_maintenance = wb.create_sheet("Phrase Maintenance")
    
    # Copy Phrase Hit Report to Phrase Maintenance
    phrase_hit_report = wb["Phrase Hit Report"]
    for row in phrase_hit_report.iter_rows(min_row=1, max_col=10, values_only=True):
        phrase_maintenance.append(row)

    # Rename columns
    new_headers = [
        'Total Hits in Reporting Period (Indexed)',
        'Count DT Changed',
        'Count Summary Changed',
        'Count DOS Changed',
        'Count Signature Changed',
        'Count Patient Found and Changed',
        'Count Provider Changed where Patient Found'
    ]
    for col, header in enumerate(new_headers, start=10):
        phrase_maintenance.cell(row=1, column=col, value=header)

    # Fill in the counted data
    for row in range(2, phrase_maintenance.max_row + 1):
        phrase = phrase_maintenance.cell(row=row, column=1).value
        if phrase in count_data:
            phrase_maintenance.cell(row=row, column=10).value = count_data[phrase]['total']
            for i in range(5):
                phrase_maintenance.cell(row=row, column=11+i).value = count_data[phrase].get(f'col_{i}', 0)
            phrase_maintenance.cell(row=row, column=16).value = count_data[phrase].get('provider_changed', 0)
        else:
            # If the phrase is not in count_data, fill with zeros
            for col in range(10, 17):
                phrase_maintenance.cell(row=row, column=col).value = 0

    # Format dates
    date_columns = ['P', 'Q', 'AA', 'AB', 'AC']
    for col in date_columns:
        for cell in ws[col]:
            cell.number_format = 'mm/dd/yyyy'

    # Set all rows to no fill
    no_fill = PatternFill(fill_type=None)
    for row in ws['A2:AG200000']:
        for cell in row:
            cell.fill = no_fill

    # Freeze pane on header row
    ws.freeze_panes = 'A2'

    # Auto fit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Set special color for added columns
    light_blue_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'AE', 'AF', 'AG']:
        ws[f'{col}1'].fill = light_blue_fill

    # Add Phrase Building sheet
    phrase_building = wb.create_sheet("Phrase Building")

    # Copy relevant columns from Documents Filed Report
    columns_to_copy = ['O', 'P', 'AG', 'AH']
    for col, source_col in enumerate(columns_to_copy, start=1):
        for row in range(1, ws.max_row + 1):
            phrase_building.cell(row=row, column=col, value=ws[f'{source_col}{row}'].value)

    # Sort the data (Note: This is a simple sort, might need adjustment for large datasets)
    data = list(phrase_building.iter_rows(values_only=True))
    headers = data.pop(0)

    print("Data in Phrase Building sheet:")
    for row in data[:10]:  # Print first 10 rows
        print(row)

    def safe_sort_key(x):
        return (
            -float(x[2]) if x[2] is not None and x[2] != '' else float('inf'),
            -float(x[3]) if x[3] is not None and x[3] != '' else float('inf')
        )

    sorted_data = sorted(data, key=safe_sort_key)
    phrase_building.delete_rows(2, phrase_building.max_row)
    phrase_building.append(headers)
    for row in sorted_data:
        phrase_building.append(row)

    # Remove duplicates
    data = list(phrase_building.iter_rows(values_only=True))
    unique_data = list(dict.fromkeys(map(tuple, data)))
    phrase_building.delete_rows(2, phrase_building.max_row)
    for row in unique_data:
        phrase_building.append(row)

    # Add Filter Updates sheet
    filter_updates = wb.create_sheet("Filter Updates")
    filter_updates['A1'] = "Version 1.26.2023"
    
    # Add text as comments (since openpyxl doesn't support text boxes)
    filter_updates['A4'].comment = openpyxl.comments.Comment(
        "Select this box to automatically apply the criteria used for Phrase Maintenance in the Documents Filed Report tab.",
        "System"
    )
    filter_updates['E4'].comment = openpyxl.comments.Comment(
        "Select this box to automatically apply the criteria used for Phrase Building to the Documents Filed Report tab.",
        "System"
    )

    filter_updates['A9'] = "Phrase Maintenance criteria: 1)Phrase is not 0. 2)Status is Manually Indexed. 3)Phrase Indexer Review = Yes."
    filter_updates['A10'] = "Phrase Building criteria: 1)Phrase is 0. 2) Status is Manually Indexed."

    # Save the workbook
    wb.save('updated_workbook.xlsx')

# Run the function
filed_documents_report_with_phrase_hit_athena()