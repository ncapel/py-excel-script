import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def count_if_multiple(ws, criteria_ranges, criteria_values):
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(row[col-1] == value for col, value in zip(criteria_ranges, criteria_values)):
            count += 1
    return count

def filed_documents_report_with_phrase_hit_athena():
    # Load the workbook
    wb = openpyxl.load_workbook('test_sheet.xlsx')
    ws = wb['Documents Filed Report']  # Assuming this is the name of your main worksheet

    # Insert columns
    ws.insert_cols(1, 6)

    # Rename columns
    headers = ['Member Match', 'Summary Match', 'DOS Match', 'Signature Match', 'Patient Match', 'Provider Match']
    for i, header in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=header)

    # Set formulas for matches
    last_row = ws.max_row
    formulas = [
        '=IF(L{0}=N{0}, "EXACTMATCH", "NEEDSREVIEW")',
        '=IF(M{0}=O{0}, "EXACTMATCH", "NEEDSREVIEW")',
        '=IF(P{0}=Q{0}, "EXACTMATCH", "NEEDSREVIEW")',
        '=IF(V{0}=W{0}, "EXACTMATCH", "NEEDSREVIEW")',
        '=IF(COUNTIF(Y{0},"*oun*"), "NOTFOUND", IF(H{0}=I{0},"EXACTMATCH","NEEDSREVIEW"))',
        '=IF(AND(E{0}="NOTFOUND",R{0}=S{0}),"PTNFEXACTMATCH",IF(AND(E{0}="NOTFOUND",R{0}<>S{0}),"PTNFNEEDSREVIEW",IF(AND(E{0}="EXACTMATCH",R{0}=S{0}),"EXACTMATCH",IF(AND(E{0}="NEEDSREVIEW",R{0}=S{0}),"EXACTMATCH","NEEDSREVIEW"))))'
    ]

    for col, formula in enumerate(formulas, start=1):
        for row in range(2, last_row + 1):
            ws.cell(row=row, column=col, value=formula.format(row))

    # Convert formulas to values
    for col in range(1, 7):
        for row in range(2, last_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = cell.value

    # Add Indexer Review column
    ws['AE1'] = 'Indexer Review'
    for row in range(2, last_row + 1):
        ws[f'AE{row}'] = f'=VLOOKUP(K{row}, \'Phrase Hit Report\'!A$2:I$30000, 5, FALSE)'

    # Convert to values
    for row in range(2, last_row + 1):
        cell = ws[f'AE{row}']
        cell.value = cell.value

    # Add Documents Manually Indexed column
    ws['AF1'] = 'Documents Manually Indexed with No Phrase by HL7 Document Type and HL7 Summary Line'
    for row in range(2, last_row + 1):
        ws[f'AF{row}'] = f'=COUNTIFS(N:N,N{row}, O:O,O{row}, X:X, "Manually Indexed", K:K,"=0")'

    # Convert to values
    for row in range(2, last_row + 1):
        cell = ws[f'AF{row}']
        cell.value = cell.value

    # Add Indexed Documents with Flag column
    ws['AG1'] = 'Indexed Documents with Flag containing No Patient Found and No Phrase Hit by HL7 Document Type and HL7 Summary Line'
    for row in range(2, last_row + 1):
        ws[f'AG{row}'] = f'=COUNTIFS(N:N,N{row},O:O,O{row},K:K,"=0",X:X,"Manually Indexed",Y:Y,"*oun*")'

    # Convert to values
    for row in range(2, last_row + 1):
        cell = ws[f'AG{row}']
        cell.value = cell.value

    # Format dates
    date_columns = ['P', 'Q', 'AA', 'AB', 'AC']
    for col in date_columns:
        for cell in ws[col]:
            cell.number_format = 'mm/dd/yyyy'

    # Set all rows to no fill
    no_fill = PatternFill(fill_type=None)
    for row in ws['A2:AG200000']:
        for cell in row:
            cell.fill = no_fill

    # Freeze pane on header row
    ws.freeze_panes = 'A2'

    # Auto fit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Set special color for added columns
    light_blue_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'AE', 'AF', 'AG']:
        ws[f'{col}1'].fill = light_blue_fill

# Add Phrase Maintenance Sheet
    phrase_maintenance = wb.create_sheet("Phrase Maintenance")
    
    # Copy Phrase Hit Report to Phrase Maintenance
    phrase_hit_report = wb["Phrase Hit Report"]
    for row in phrase_hit_report.iter_rows(min_row=1, max_col=10, values_only=True):
        phrase_maintenance.append(row)

    # Rename columns
    new_headers = [
        'Total Hits in Reporting Period (Indexed)',
        'Count DT Changed',
        'Count Summary Changed',
        'Count DOS Changed',
        'Count Signature Changed',
        'Count Patient Found and Changed',
        'Count Provider Changed where Patient Found'
    ]
    for col, header in enumerate(new_headers, start=10):
        phrase_maintenance.cell(row=1, column=col, value=header)

    # Generate values for new columns using Python functions
    for row in range(2, phrase_maintenance.max_row + 1):
        phrase_value = phrase_maintenance.cell(row=row, column=1).value
        
        # Total Hits in Reporting Period (Indexed)
        phrase_maintenance.cell(row=row, column=10).value = count_if_multiple(
            ws, [12, 25, 32], [phrase_value, "Manually Indexed", "Yes"])
        
        # Count DT Changed
        phrase_maintenance.cell(row=row, column=11).value = count_if_multiple(
            ws, [12, 25, 32, 1], [phrase_value, "Manually Indexed", "Yes", "NEEDSREVIEW"])
        
        # Count Summary Changed
        phrase_maintenance.cell(row=row, column=12).value = count_if_multiple(
            ws, [12, 25, 32, 2], [phrase_value, "Manually Indexed", "Yes", "NEEDSREVIEW"])
        
        # Count DOS Changed
        phrase_maintenance.cell(row=row, column=13).value = count_if_multiple(
            ws, [12, 25, 32, 3], [phrase_value, "Manually Indexed", "Yes", "NEEDSREVIEW"])
        
        # Count Signature Changed
        phrase_maintenance.cell(row=row, column=14).value = count_if_multiple(
            ws, [12, 25, 32, 4], [phrase_value, "Manually Indexed", "Yes", "NEEDSREVIEW"])
        
        # Count Patient Found and Changed
        phrase_maintenance.cell(row=row, column=15).value = count_if_multiple(
            ws, [12, 25, 32, 5], [phrase_value, "Manually Indexed", "Yes", "NEEDSREVIEW"])
        
        # Count Provider Changed where Patient Found
        phrase_maintenance.cell(row=row, column=16).value = count_if_multiple(
            ws, [12, 25, 32, 5, 6], [phrase_value, "Manually Indexed", "Yes", "EXACTMATCH", "NEEDSREVIEW"])

    # Filter for Indexer Review = YES
    phrase_maintenance.auto_filter.ref = phrase_maintenance.dimensions
    phrase_maintenance.auto_filter.add_filter_column(4, ["Yes"])

    # Add Phrase Building sheet
    phrase_building = wb.create_sheet("Phrase Building")

    # Copy relevant columns from Documents Filed Report
    columns_to_copy = ['O', 'P', 'AG', 'AH']
    for col, source_col in enumerate(columns_to_copy, start=1):
        for row in range(1, ws.max_row + 1):
            phrase_building.cell(row=row, column=col, value=ws[f'{source_col}{row}'].value)

    # Sort the data
    phrase_building.auto_filter.ref = phrase_building.dimensions
    phrase_building.auto_filter.add_sort_condition("C2:C1048576", descending=True)
    phrase_building.auto_filter.add_sort_condition("D2:D1048576", descending=True)

    # Remove duplicates
    data = [list(row) for row in phrase_building.iter_rows(values_only=True)]
    unique_data = list(set(map(tuple, data)))
    phrase_building.delete_rows(2, phrase_building.max_row)
    for row in unique_data:
        phrase_building.append(row)

    # Add Filter Updates sheet
    filter_updates = wb.create_sheet("Filter Updates")
    filter_updates['A1'] = "Version 1.26.2023"
    
    # Add text boxes (Note: openpyxl doesn't support adding shapes directly, so we'll add comments instead)
    filter_updates['A4'].comment = openpyxl.comments.Comment(
        "Select this box to automatically apply the criteria used for Phrase Maintenance in the Documents Filed Report tab.",
        "System"
    )
    filter_updates['E4'].comment = openpyxl.comments.Comment(
        "Select this box to automatically apply the criteria used for Phrase Building to the Documents Filed Report tab.",
        "System"
    )

    filter_updates['A9'] = "Phrase Maintenance criteria: 1)Phrase is not 0. 2)Status is Manually Indexed. 3)Phrase Indexer Review = Yes."
    filter_updates['A10'] = "Phrase Building criteria: 1)Phrase is 0. 2) Status is Manually Indexed."

    # Save the workbook
    wb.save('updated_workbook.xlsx')

# Run the function
filed_documents_report_with_phrase_hit_athena()